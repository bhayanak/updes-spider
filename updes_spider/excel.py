"""Export parsed tables to Excel workbooks."""

from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .parse import DataTable

_HEADER_FILL = PatternFill("solid", fgColor="8080FF")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_WRAP = Alignment(wrap_text=True, vertical="top")


def write_table_workbook(
    path: Path,
    tables: List[DataTable],
    *,
    meta: List[tuple] | None = None,
) -> None:
    """Write one workbook per table action.

    Multiple data tables (e.g. Table 2A / 2B) become separate sheets.
    ``meta`` is a list of (label, value) rows written atop the first sheet.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    if not tables:
        ws = wb.create_sheet("data")
        ws["A1"] = "No data"
    for i, dt in enumerate(tables, start=1):
        ws = wb.create_sheet(f"Table {i}" if len(tables) > 1 else "data")
        r = 1
        if i == 1 and meta:
            for label, value in meta:
                ws.cell(row=r, column=1, value=label).font = Font(bold=True)
                ws.cell(row=r, column=2, value=value)
                r += 1
            r += 1  # blank spacer row
        start = r
        for row in dt.rows:
            for c, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.alignment = _WRAP
            r += 1
        # style the first data row as a header
        if dt.rows:
            for c in range(1, len(dt.rows[0]) + 1):
                cell = ws.cell(row=start, column=c)
                cell.fill = _HEADER_FILL
                cell.font = _HEADER_FONT
        _autosize(ws, dt.rows)

    wb.save(path)


def _autosize(ws, rows: List[List[str]]) -> None:
    if not rows:
        return
    widths: dict = {}
    for row in rows:
        for c, val in enumerate(row, start=1):
            ln = min(len(str(val)), 60)
            if ln > widths.get(c, 0):
                widths[c] = ln
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = max(8, w + 2)
